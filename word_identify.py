# -*- coding: utf-8 -*-
"""
Created on Wed Aug  8 06:33:36 2018

@author: useu
"""
from tkinter import ttk

import tkinter as tk
import tkinter
import tkinter.messagebox
import tkinter.font as tkFont
import random 
from openpyxl import load_workbook
from win32api import GetSystemMetrics
size = []
word = []
testword = []
finallist = []
size1 = []
kind =[]
text = []
var1 =[]
errorlist = []
check = []
num_error = []
list2 = []
correct1 = []
text_list = []

class myWindow:
    def __init__(self, win, flag):
     
     self.top = tkinter.Toplevel(win, width=size1[0], height=size1[1],bg = '#000000')
     self.top.title('test')
     self.top.attributes('-fullscreen', True) 
     self.top.config(cursor="none")
     
     if flag == 1:   
        count = ['0']
        fontsize=['-10']
        num = ['0']
        pause_list = ['0']     
        
        if word[0] == '1':  
            ft = tkFont.Font(family = 'NotoSerifCJKtc-Black',size = fontsize[0])            
            test = dict[list2[-1]]      
            label2 = tkinter.Label(self.top,text =test[0], font=ft,fg ='#0000ff', bg = '#000000',anchor='center')            
        elif word[0] == '2':
            ft = tkFont.Font(family = 'NotoSerifCJKtc-Regular',size = fontsize[0])
            test = dict[list2[-1]] 
            label2 = tkinter.Label(self.top,text = test[0],font=ft,fg ='#0000ff', bg = '#000000',anchor='center')
        elif word[0] == '3':
            ft = tkFont.Font(family = 'NotoSerifCJKtc-Black',size = fontsize[0])
            test = dict[list2[-1]]  
            label2 = tkinter.Label(self.top,text = test[0],font=ft,fg ='#ffffff', bg = '#0000ff',anchor='center')
        elif word[0] == '4':
            ft = tkFont.Font(family = 'NotoSerifCJKtc-Regular',size = fontsize[0])
            test = dict[list2[-1]]
            label2 = tkinter.Label(self.top,text = test[0],font=ft,fg ='#ffffff', bg = '#0000ff', anchor='center')
        elif word[0] == '5':
            ft = tkFont.Font(family = 'NotoSansCJKtc-Black',size = fontsize[0] )
            test = dict[list2[-1]]   
            label2 = tkinter.Label(self.top,text = test[0],font=ft,fg ='#0000ff', bg = '#000000', anchor='center')
        elif word[0] == '6':
            ft = tkFont.Font(family = 'NotoSansCJKtc-Regular',size =  fontsize[0])
            test = dict[list2[-1]]
            label2 = tkinter.Label(self.top,text = test[0],font=ft,fg ='#0000ff', bg = '#000000', anchor='center')
        elif word[0] == '7':
            ft = tkFont.Font(family = 'NotoSansCJKtc-Black',size =  fontsize[0])
            test = dict[list2[-1]] 
            label2 = tkinter.Label(self.top,text = test[0],font=ft,fg ='#ffffff', bg = '#0000ff', anchor='center')
        elif word[0] == '8-1':
            ft = tkFont.Font(family = 'NotoSansCJKtc-Regular',size =  fontsize[0])
            test = dict[list2[-1]] 
            label2 = tkinter.Label(self.top,text = test[0],font=ft,fg ='#ffffff', bg = '#0000ff', anchor='center')
        elif word[0] == '8-2':
            ft = tkFont.Font(family = 'NotoSansCJKtc-Regular',size =  fontsize[0])
            test = dict[list2[-1]]
            label2 = tkinter.Label(self.top,text = test[0],font=ft,fg ='#ffffff', bg = '#0000ff', anchor='center')
        else:
            ft = tkFont.Font(family = 'NotoSansCJKtc-Regular',size =  fontsize[0])
            test = dict[list2[-1]]  
            label2 = tkinter.Label(self.top,text = test[0],font=ft,fg ='#ffffff', bg = '#0000ff', anchor='center')
                                   
        def change(self):
            if pause_list[0] == '0':
                mid = []
                text.append(test[int(num[0])])
                size.append(fontsize[0])
                kind.append(word[0])
                text_list.append(list2[-1])
                if len(correct1)< len(size):
                    correct1.append(fontsize[0])
                else:
                    pass
                    
                if test[int(num[0])] == test[29]:
                    label2.config(text= 'NEXT',font = ft)
                else:
                    pass
                num_error.append(count[0])
                if  count[0] == '0':
                    check.append("")
                else:
                    for item in errorlist:
                       mid.append(item)
                    check.append(mid)
                u = 0
                count.clear()
                count.append(u)
            
                errorlist.clear()
                d = str(int(num[0])+1)
                            
                num.clear()
                num.append(d)           
                fontsize.clear()
                h=-10
                fontsize.append(h)
                ft.configure(size= int(h))
                label2.config(text= test[int(d)],font = ft)
            else:
                 pass
            
        
        def error(self):
            o = str(int(count[0])+1)
            count.remove(count[0])
            errorlist.append('錯誤{},大小{}'.format(o,fontsize[0]))
            count.append(o)
            
        def zoom(self):
            #print(pause_list)
            if pause_list[0] == '0':
                print(fontsize[0])
                z = str(int(fontsize[0])-1)
                fontsize.clear()
                fontsize.append(z)
                ft.configure(size=int(z))            
                
                #label2.place(x=size1[0]/2, y=size1[1]/2,anchor='center')
            else:
                pass
        def pause(self):
            if pause_list[0] == '0':
                pause_list.clear()
                pause_list.append('1')
                print(pause_list)
            else:
                pause_list.clear()
                pause_list.append('0')
                print(pause_list)
        def correct(self):
            if len(correct1) == len(size)+1:
                pass
            else:
                correct1.append(fontsize[0])
        label2.bind('<Pause>',pause)    
        label2.bind('<KeyPress-+>',zoom)
        label2.bind('<KeyPress-space>',change)
        label2.bind('<Alt_L>',correct)
        label2.bind('<Control_L>',error)
        label2.bind('<Escape>',lambda e:self.top.destroy())
        label2.focus_set()
        label2.place(x=size1[0]/2, y=size1[1]/2,anchor='center')
     elif flag ==2:
        fontsize=['10']
        ft = tkFont.Font(family = '標楷體',size = fontsize[0])
        label2 = tkinter.Label(self.top,text = '測',font=ft,fg ='#ffffff', bg = '#000000', anchor='center')
        def change1(self):
            fontsize.clear()
            h=10
            fontsize.append(h)
            ft.config(size= int(h))
            label2.config(text= '龘',font = ft, bg = '#0000ff')
            pass
        def zoom1(self):
            z = str(int(fontsize[0])+1)
            fontsize.clear()
            fontsize.append(z)
            ft.config(size= int(z))
            
        label2.bind('<KeyPress-+>',zoom1)
        label2.bind('<KeyPress-space>',change1)
        label2.bind('<Escape>',lambda e:self.top.destroy())
        label2.focus_set()
        label2.place(x=size1[0]/2, y=size1[1]/2, anchor= "center")
        
  
win = tkinter.Tk()
width = GetSystemMetrics(0)
height = GetSystemMetrics(1)
size1.append(width)
size1.append(height)
win.geometry("650x400")
window1 = tkinter.IntVar(win, value=0)
window2 = tkinter.IntVar(win, value=0)
win.title('main window')
def record():
            book = load_workbook('D://python/word_identify/test.xlsx')
            #book_w = copy(book_r)
            #book.get_sheet_by_name(book.sheetnames[0])
            a = int(var1[0])
            print(a)
            sheet_1 = book.get_sheet_by_name(book.sheetnames[a])
            num =2
            num1 =2
            num2 =2
            num3 =2
            num4 =2
            num5 = 2
            num6 = 2
            print(text_list)
            print(kind)
            print(text)
            print(size)
            print(check)
            print(num_error)
            for item in text_list:
                sheet_1.cell(num, 1, item)
                num +=1
            for item in kind:
                sheet_1.cell(num1,2, item)
                num1 +=1
            for item in text:
                sheet_1.cell(num2,3, item)
                num2 +=1
            for item in correct1:
                sheet_1.cell(num3,4, abs(int(item)))
                num3 +=1
            for item in size:
                sheet_1.cell(num4,5, abs(int(item)))
                num4 +=1
            for item in num_error:
                sheet_1.cell(num5,6, int(item))
                num5 +=1
            for item in check:
                r = 7
                for x in item:
                    sheet_1.cell(num6,r,x)
                    r+=1
                num6 +=1
            book.save('D://python/word_identify/test.xlsx')
def sure():
    var = entry1.get()
    var1.append(var)
    #print(var1[0])
    #print(type(int(var1[0])))
    book = load_workbook('D://python/word_identify/test.xlsx')
    #book_w = copy(book)    
    sheet1 = book.create_sheet('test{}'.format(int(var)))
    sheet1.cell(1,1, '字庫')
    sheet1.cell(1,2, '樣式')
    sheet1.cell(1,3, '文字')
    sheet1.cell(1,4, '量測結果(px)')
    sheet1.cell(1,5,'舒適大小')
    sheet1.cell(1,6, '錯誤')
    sheet1.cell(1,7, '附註')
    book.save('D://python/word_identify/test.xlsx')
    entry2.insert('insert','工作表建立成功!!')
    print('工作表建立成功!!')


def buttonClick1():
    if window1.get()==0:
        window1.set(1)
        w1 = myWindow(win, 1)
        button1.wait_window(w1.top)
        window1.set(0)

def buttonClick2():
    if window2.get()==0:
        window2.set(1)
        w1 = myWindow(win,2)
        button5.wait_window(w1.top)
    

f = open("D://python/word_identify/text.txt",'r')   #以文字檔為資料庫
ll = f.readlines()
for item in ll:
    testword.append(item[0].strip())
    #random.shuffle(testword)
#print(testword)
global textlistA,textlistB,textlistC,textlistD,textlistE,textlistF,textlistG,textlistH,textlistI,textlistJ
textlistA=testword[:30]
textlistB=testword[30:60]
textlistC=testword[60:90]
textlistD=testword[90:120]
textlistE=testword[120:150]
textlistF=testword[150:180]
textlistG=testword[180:210]
textlistH=testword[210:240]
textlistI=testword[240:270]
textlistJ=testword[270:300]

random.shuffle(textlistA)
random.shuffle(textlistB)
random.shuffle(textlistC)
random.shuffle(textlistD)
random.shuffle(textlistE)
random.shuffle(textlistF)
random.shuffle(textlistG)
random.shuffle(textlistH)
random.shuffle(textlistI)
random.shuffle(textlistJ)
dict = {'A': textlistA,'B': textlistB,'C': textlistC,'D': textlistD,'E': textlistE,'F': textlistF,'G': textlistG,'H': textlistH,'I': textlistI,'J': textlistJ}
def choose1():
    j = choice1.get()
    print(j)
    if j in word:
        pass
    else:
        word.append(j)
    #print(word)
       
#def reset():
    #word.remove(word[0]) 

def clearall():
    word.clear()          
def go(*arg):
    #print(comboxlist.get())
    list2.append(comboxlist.get())
    print(list2)
    word.clear()
frame1 = tk.Frame(win)
frame1.pack(side='top')
label3=tk.Label(frame1, text="受測者：")
entry1 = tk.Entry(frame1)
label3.grid(row=0, column=0)
entry1.grid(row=0, column=1)
button1 = tk.Button(frame1, text="建立工作表",command=sure)
button1.grid(row=0, column=10)

entry2 = tk.Entry(frame1)
entry2.grid(row=10, column=1)

choice = tk.StringVar()
choice1 = tk.StringVar()
label = tk.Label(win, text="選擇詞庫：")
label.pack()
label.place(x=0,y=80)
comvalue=tkinter.StringVar()#窗体自带的文本，新建一个值
comboxlist=ttk.Combobox(win,textvariable=comvalue) #初始化
comboxlist["values"]=("A","B","C","D","E","F","G","H","I","J")
comboxlist.current(0)  #选择第一个
comboxlist.bind("<<ComboboxSelected>>",go)  #绑定事件,(下拉列表框被选中时，绑定go()函数)
comboxlist.pack()
comboxlist.place(x=10,y=100)

label1 = tk.Label(win, text="選擇字體樣式：")
label1.pack()
label1.place(x=200,y=80)
item4 = tk.Radiobutton(win, text="樣式1", value="1", variable=choice1, command=choose1)
item4.pack()
item4.place(x=230,y=120)
item5 = tk.Radiobutton(win, text="樣式2", value="2", variable=choice1, command=choose1)
item5.pack()
item5.place(x=230,y=140)
item6 = tk.Radiobutton(win, text="樣式3", value="3", variable=choice1, command=choose1)
item6.pack()
item6.place(x=230,y=160)
item7 = tk.Radiobutton(win, text="樣式4", value="4", variable=choice1, command=choose1)
item7.pack()
item7.place(x=230,y=180)
item8 = tk.Radiobutton(win, text="樣式5", value="5", variable=choice1, command=choose1)
item8.pack()
item8.place(x=230,y=200)
item9 = tk.Radiobutton(win, text="樣式6", value="6", variable=choice1, command=choose1)
item9.pack()
item9.place(x=350,y=120)
item10 = tk.Radiobutton(win, text="樣式7", value="7", variable=choice1, command=choose1)
item10.pack()
item10.place(x=350,y=140)
item11 = tk.Radiobutton(win, text="樣式8-1", value="8-1", variable=choice1, command=choose1)
item11.pack()
item11.place(x=350,y=160)
item11 = tk.Radiobutton(win, text="樣式8-2", value="8-2", variable=choice1, command=choose1)
item11.pack()
item11.place(x=350,y=180)
item11 = tk.Radiobutton(win, text="樣式8-3", value="8-3", variable=choice1, command=choose1)
item11.pack()
item11.place(x=350,y=200)



canvas1=tk.Canvas(win,bg='black',height=200,width=1)
canvas1.pack()
canvas1.place(x=180,y=80)
canvas2=tk.Canvas(win,bg='black',height=200,width=1)
canvas2.pack()
canvas2.place(x=450,y=80)
button1 = tkinter.Button(win, text='開始測驗', command=buttonClick1,  width=15, height=2)
button1.place(x=500,y=80)
#button2 = tkinter.Button(win, text='重設字體', command=reset)
#button2.place(x=380,y=230)
button3 = tkinter.Button(win, text='清除', command=clearall,width=15, height=2)
button3.place(x=500,y=150)
button4 = tkinter.Button(win, text='紀錄', command=record, width=15, height=2)
button4.place(x=500,y=220)
button5 = tk.Button(win, text="範例", command=buttonClick2)
button5.place(x=470,y=0)
win.mainloop()